from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime
from pathlib import Path
from pprint import pprint
from dataclasses import make_dataclass

import tennis_tournamend_drawing as drawing
import settings

def create_starting_positions(tournament_size):
    seed_positions = {
        8: [1, 8],
        16: [1, 5, 12, 16], 
        32: [1, 8, 9, 16, 17, 24, 25, 32]
    }
    acctual_row = 8
    second_round_row = 9
    starting_positions = {}
    PositionsAttributes = make_dataclass(
       "PlayersAttributes", 
       ["rank_field", "name_field", "seed_field", "lp_field", 
       "second_round_field"]
    )
    for i in range(1, tournament_size + 1):
        key = i 
        starting_positions[key] = PositionsAttributes(
            rank_field=f"C{acctual_row}",
            name_field=f"E{acctual_row}",
            seed_field=None,
            lp_field=None,
            second_round_field=f"H{second_round_row}",
        )
        if i % 2 == 0:
            second_round_row += 4
        if key in seed_positions[tournament_size]:
            starting_positions[key].seed_field = f"D{acctual_row}"
            starting_positions[key].lp_field = f"A{acctual_row}"
        acctual_row += 2
    return starting_positions


def read_excel(tournament_size):
    path = (
        settings.paths["torunamets_samples_path"]
        / f"cup{tournament_size}.xlsx"
    )
    wb = load_workbook(filename=path)
    return wb


def save_excel(excel_to_save, tournament_name):
    path = (
        settings.paths["created_excel_files_path"]
        / f"{datetime.now().strftime('%d.%m.%Y')} {tournament_name}.xlsx"
    )
    return excel_to_save.save(path)


def tournament_information_fill(
    excel_sheet, tournament_name, tournament_arbiter, tournament_category,
    tournament_city, tournament_date, tournament_drawing_date, 
    draw_witnesses_1, draw_witnesses_2, seed_list, tournament_size
):
    excel_sheet["C3"] = tournament_name
    excel_sheet["C4"] = tournament_category
    excel_sheet["H1"] = tournament_arbiter
    excel_sheet["H2"] = tournament_category
    excel_sheet["H3"] = tournament_city
    excel_sheet["H4"] = tournament_date
    if tournament_size == 32:
        excel_sheet["D73"] = tournament_drawing_date   
        excel_sheet["E76"] = draw_witnesses_1
        excel_sheet["E77"] = draw_witnesses_2
        bottom_seed_rows = list(range(73, 81))
    elif tournament_size == 16:
        excel_sheet["D59"] = tournament_drawing_date   
        excel_sheet["E62"] = draw_witnesses_1
        excel_sheet["E63"] = draw_witnesses_2
        bottom_seed_rows = list(range(59, 63))
    for c, v in enumerate(seed_list):
        excel_sheet[f"L{bottom_seed_rows[c]}"] = v


def bold_font(excel_position):
    excel_position.font = Font(bold=True)


def write_excel(
    tournament_name, tournament_category, tournament_arbiter, tournament_city,
    tournament_date, tournament_size, tournament_drawing_date, 
    draw_witnesses_1, draw_witnesses_2
):
    excel_posiotions = create_starting_positions(tournament_size)
    wb = read_excel(tournament_size)
    ws = wb.active
    seed_players = []
    for players in drawing.players_placement.values():
        for key, value in players.items():
            ws[excel_posiotions[key].name_field] = value.name
            if value.rank > 0:
                ws[excel_posiotions[key].rank_field] = value.rank
            if value.second_round > 0:
                last_name = value.name.split()[0]
                ws[excel_posiotions[key].second_round_field] = last_name
                if value.seed > 0:
                    bold_font(ws[excel_posiotions[key].second_round_field])
            if value.seed > 0:
                seed_players.append(value.name)
                ws[excel_posiotions[key].seed_field] = value.seed
                bold_font(ws[excel_posiotions[key].name_field])
                bold_font(ws[excel_posiotions[key].rank_field])
                bold_font(ws[excel_posiotions[key].seed_field])
                bold_font(ws[excel_posiotions[key].lp_field])
    tournament_information_fill(
        excel_sheet=ws, 
        tournament_name=tournament_name,
        tournament_arbiter=tournament_arbiter, 
        tournament_category=tournament_category,
        tournament_city=tournament_city,
        tournament_date=tournament_date,
        tournament_drawing_date=tournament_drawing_date, 
        draw_witnesses_1=draw_witnesses_1,
        draw_witnesses_2=draw_witnesses_2, 
        seed_list=seed_players,
        tournament_size=tournament_size
    )
    return save_excel(excel_to_save=wb, tournament_name=tournament_name)

# SETTINGS
write_excel(
    tournament_name=settings.config["tournament_name"],
    tournament_category=settings.config["tournament_category"],
    tournament_arbiter=settings.config["tournament_arbiter"],
    tournament_city=settings.config["tournament_city"],
    tournament_date=settings.config["tournament_date"],
    tournament_size=settings.config["tournament_size"],
    tournament_drawing_date=settings.config["tournament_drawing_date"],
    draw_witnesses_1=settings.config["draw_witnesses_1"],
    draw_witnesses_2=settings.config["draw_witnesses_2"],
)
